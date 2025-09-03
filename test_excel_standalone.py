#!/usr/bin/env python3
"""
Teste standalone para verificar se a lógica do Excel funciona corretamente
"""

import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def test_excel_generation_standalone():
    """Testa a geração de arquivo Excel com dados de exemplo - versão standalone"""
    
    # Dados de teste
    test_data = {
        "fichas_tecnicas": [
            {
                "nome_preparacao": "Bolo de Chocolate",
                "rendimento_porcoes": 10,
                "preco_venda": 25.00,
                "ingredientes": [
                    {
                        "nome": "Farinha de Trigo",
                        "unidade": "kg",
                        "quantidade": 0.5,
                        "fator_correcao": 1.0,
                        "custo_unitario": 4.50
                    },
                    {
                        "nome": "Açúcar",
                        "unidade": "kg", 
                        "quantidade": 0.3,
                        "fator_correcao": 1.0,
                        "custo_unitario": 3.20
                    },
                    {
                        "nome": "Chocolate em Pó",
                        "unidade": "kg",
                        "quantidade": 0.2,
                        "fator_correcao": 1.0,
                        "custo_unitario": 15.00
                    }
                ],
                "modo_preparo": [
                    "Misture os ingredientes secos",
                    "Adicione os líquidos gradualmente",
                    "Asse em forno pré-aquecido a 180°C por 40 minutos"
                ]
            }
        ],
        "base_de_insumos": [
            {
                "ingrediente": "Farinha de Trigo",
                "unidade": "kg",
                "preco": 4.50,
                "fator_correcao": 1.0,
                "fornecedor": "Fornecedor A",
                "data_cotacao": "2024-01-15"
            },
            {
                "ingrediente": "Açúcar",
                "unidade": "kg",
                "preco": 3.20,
                "fator_correcao": 1.0,
                "fornecedor": "Fornecedor B", 
                "data_cotacao": "2024-01-15"
            },
            {
                "ingrediente": "Chocolate em Pó",
                "unidade": "kg",
                "preco": 15.00,
                "fator_correcao": 1.0,
                "fornecedor": "Fornecedor C",
                "data_cotacao": "2024-01-15"
            }
        ]
    }
    
    print("🧪 Testando geração de arquivo Excel (standalone)...")
    print(f"📊 Dados: {len(test_data['fichas_tecnicas'])} fichas, {len(test_data['base_de_insumos'])} insumos")
    
    try:
        # Criar workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        fichas = test_data.get('fichas_tecnicas', [])
        insumos = test_data.get('base_de_insumos', [])
        
        # Estilos
        font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        fill_titulo = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        font_subtitulo = Font(name='Arial', size=11, bold=True)
        font_bold = Font(name='Arial', size=10, bold=True)
        fill_header_tabela = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Criar aba para ficha técnica
        for idx, ficha in enumerate(fichas, 1):
            sheet_name = f"Ficha_{idx}_{ficha.get('nome_preparacao', 'Receita')[:20]}"
            sheet = workbook.create_sheet(title=sheet_name)

            # Cabeçalho
            sheet.merge_cells('A1:G1')
            cell = sheet['A1']
            cell.value = "FICHA TÉCNICA DE PREPARAÇÃO"
            cell.font = font_titulo
            cell.fill = fill_titulo
            cell.alignment = align_center
            sheet.row_dimensions[1].height = 25
            
            sheet['A3'] = "NOME DA PREPARAÇÃO:"
            sheet['A3'].font = font_subtitulo
            sheet.merge_cells('B3:G3')
            sheet['B3'] = ficha.get("nome_preparacao", "").upper()
            sheet['B3'].font = font_subtitulo
            
            # Headers da tabela de ingredientes
            headers_tabela = ["INGREDIENTES", "UNIDADE", "QUANTIDADE", "FATOR CORREÇÃO", "PESO CORRIGIDO", "CUSTO UNITÁRIO", "CUSTO TOTAL"]
            for col_num, header_text in enumerate(headers_tabela, 1):
                cell = sheet.cell(row=7, column=col_num, value=header_text)
                cell.font = font_bold
                cell.fill = fill_header_tabela
                cell.border = border_thin
            
            # Dados dos ingredientes
            linha_atual = 8
            for ingrediente in ficha.get("ingredientes", []):
                nome = ingrediente.get('nome', 'N/A')
                unidade = ingrediente.get('unidade', 'N/A')
                quantidade_original = ingrediente.get('quantidade', 0)
                fator_correcao = ingrediente.get('fator_correcao', 1.0)
                custo_unitario = ingrediente.get('custo_unitario', 0)
                quantidade_corrigida = quantidade_original * fator_correcao

                sheet.cell(row=linha_atual, column=1, value=nome)
                sheet.cell(row=linha_atual, column=2, value=unidade)
                sheet.cell(row=linha_atual, column=3, value=quantidade_original)
                sheet.cell(row=linha_atual, column=4, value=fator_correcao)
                sheet.cell(row=linha_atual, column=5, value=quantidade_corrigida)
                sheet.cell(row=linha_atual, column=6, value=custo_unitario)
                sheet.cell(row=linha_atual, column=7, value=f'=E{linha_atual}*F{linha_atual}')
                
                for col in range(1, 8):
                    cell = sheet.cell(row=linha_atual, column=col)
                    cell.border = border_thin
                    if col in [3, 4, 5]:
                        cell.number_format = '0.000'
                    if col in [6, 7]:
                        cell.number_format = 'R$ #,##0.00'
                linha_atual += 1

            # Ajustar larguras das colunas
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 15
        
        # Criar aba de insumos
        if insumos:
            sheet_base = workbook.create_sheet(title="Base de Insumos")
            
            headers_base = ["INGREDIENTE", "UNIDADE", "PREÇO UNITÁRIO", "FATOR CORREÇÃO", "FORNECEDOR", "DATA DE COTAÇÃO"]
            widths = [30, 12, 15, 15, 25, 18]
            
            for col, (header, width) in enumerate(zip(headers_base, widths), 1):
                cell = sheet_base.cell(row=1, column=col, value=header)
                cell.font = font_bold
                cell.fill = fill_header_tabela
                cell.border = border_thin
                sheet_base.column_dimensions[get_column_letter(col)].width = width
            
            for row, item in enumerate(insumos, 2):
                sheet_base.cell(row=row, column=1, value=item.get("ingrediente"))
                sheet_base.cell(row=row, column=2, value=item.get("unidade"))
                
                preco_cell = sheet_base.cell(row=row, column=3, value=item.get("preco"))
                preco_cell.number_format = 'R$ #,##0.00'
                
                fator_cell = sheet_base.cell(row=row, column=4, value=item.get("fator_correcao", 1.0))
                fator_cell.number_format = '0.000'
                
                sheet_base.cell(row=row, column=5, value=item.get("fornecedor"))
                sheet_base.cell(row=row, column=6, value=item.get("data_cotacao"))
                
                for col in range(1, 7):
                    sheet_base.cell(row=row, column=col).border = border_thin
        
        # Salvar arquivo
        output_dir = "/Users/cairorocha/Documents/fichas_tecnicas1/output"
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"TESTE_FICHA_TECNICA_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        workbook.save(filepath)
        print(f"✅ Arquivo salvo: {filepath}")
        
        # Verificar arquivo
        if os.path.exists(filepath):
            size = os.path.getsize(filepath) / 1024  # KB
            print(f"📁 Tamanho: {size:.1f} KB")
            
            # Tentar abrir para verificar validade
            wb = openpyxl.load_workbook(filepath)
            print(f"📋 Planilhas: {[ws.title for ws in wb.worksheets]}")
            
            # Verificar conteúdo da primeira planilha
            first_sheet = wb.worksheets[0]
            print(f"📄 Primeira célula: {first_sheet.cell(1, 1).value}")
            print(f"📄 Nome da preparação: {first_sheet.cell(3, 2).value}")
            
            wb.close()
            print("✅ Arquivo Excel válido!")
            return True
        else:
            print("❌ Arquivo não foi criado")
            return False
            
    except Exception as e:
        print(f"❌ Erro: {e}")
        return False

if __name__ == "__main__":
    success = test_excel_generation_standalone()
    if success:
        print("\n🎉 TESTE PASSOU - O gerador Excel funciona corretamente!")
    else:
        print("\n💥 TESTE FALHOU - Há problemas na geração do Excel")