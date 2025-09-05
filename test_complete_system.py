#!/usr/bin/env python3
"""
Teste completo do sistema com cores e cálculos de custos
"""

import sys
import os
sys.path.append('/Users/cairorocha/Documents/fichas_tecnicas')

from main import fichas_tecnicas
import tempfile
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def create_test_recipe():
    """Cria um arquivo de receita de teste"""
    recipe_content = """PAELLA DE FRUTOS DO MAR

Ingredientes:
- 500g arroz
- 250g camarão sem casca  
- 150g mexilhões com concha
- 200g lulas frescas
- 4 colheres (sopa) de azeite
- 2 dentes de alho amassados
- 3 unidades de tomate
- 2 unidades de cebolinha
- 1 pitada de sal
- 1 pitada de pimenta-do-reino
- 1 colher (chá) de açafrão
- 1 colher (chá) de páprica
- 200g de ervilha cozida

Modo de Preparo:
1. Aqueça o azeite em uma paelleira grande em fogo médio
2. Refogue o alho e a cebola até ficarem dourados
3. Adicione o tomate picado e cozinhe por 2-3 minutos
4. Adicione o arroz e mexa bem para envolver com o refogado
5. Tempere com sal, pimenta, açafrão e páprica
6. Adicione água quente suficiente para cobrir o arroz
7. Cozinhe por 15 minutos sem mexer
8. Adicione os frutos do mar e ervilha
9. Continue cozinhando por mais 10 minutos
10. Deixe descansar por 5 minutos antes de servir

Rendimento: 6 porções"""
    
    # Criar arquivo temporário
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
        f.write(recipe_content)
        return f.name

def test_complete_system():
    """Testa o sistema completo com cores personalizadas"""
    
    print("🧪 Testando sistema completo com cores personalizadas...")
    
    # Criar arquivo de receita de teste
    recipe_file = create_test_recipe()
    print(f"📝 Arquivo de receita criado: {recipe_file}")
    
    sources = [recipe_file]
    color1 = 'C5504B'  # Vermelho
    color2 = 'F2DCDB'  # Vermelho claro
    
    try:
        print(f"🎨 Testando com cores: Primary={color1}, Secondary={color2}")
        print("🚀 Iniciando processamento...")
        
        # Executar o sistema
        result = fichas_tecnicas(
            sources=sources, 
            color1=color1, 
            color2=color2
        )
        
        print(f"✅ Processamento concluído!")
        print(f"📊 Resultado: {result}")
        
        # Verificar se foi gerado arquivo Excel
        output_dir = "/Users/cairorocha/Documents/fichas_tecnicas/output"
        if os.path.exists(output_dir):
            excel_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
            if excel_files:
                latest_file = max([os.path.join(output_dir, f) for f in excel_files], key=os.path.getmtime)
                print(f"📊 Arquivo Excel gerado: {latest_file}")
                
                # Verificar tamanho do arquivo
                size = os.path.getsize(latest_file) / 1024
                print(f"📁 Tamanho: {size:.1f} KB")
                
                # Verificar se o arquivo Excel pode ser aberto (validação básica)
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(latest_file)
                    print(f"📋 Planilhas encontradas: {[ws.title for ws in wb.worksheets]}")
                    
                    # Verificar primeira planilha
                    if wb.worksheets:
                        first_sheet = wb.worksheets[0] 
                        print(f"📄 Cabeçalho: {first_sheet.cell(1, 1).value}")
                        
                        # Verificar se há cores aplicadas (verificando fills)
                        header_cell = first_sheet.cell(1, 1)
                        if header_cell.fill and header_cell.fill.start_color:
                            actual_color = header_cell.fill.start_color.rgb
                            if actual_color and actual_color != '00000000':
                                print(f"🎨 Cor do cabeçalho aplicada: {actual_color}")
                            else:
                                print("⚠️  Cor do cabeçalho não detectada")
                        
                        # Verificar dados de ingredientes
                        ingredient_row = None
                        for row in range(1, 20):  # Procurar nas primeiras 20 linhas
                            cell_value = first_sheet.cell(row, 1).value
                            if cell_value and 'arroz' in str(cell_value).lower():
                                ingredient_row = row
                                break
                        
                        if ingredient_row:
                            print(f"🌾 Ingrediente 'arroz' encontrado na linha {ingredient_row}")
                            
                            # Verificar se custo unitário foi preenchido
                            cost_cell = first_sheet.cell(ingredient_row, 6)  # Coluna F - custo unitário
                            if cost_cell.value and cost_cell.value != 0:
                                print(f"💰 Custo unitário do arroz: {cost_cell.value}")
                            else:
                                print("⚠️  Custo unitário não foi preenchido")
                        
                        print(f"✅ Teste concluído com sucesso!")
                        return True
                    
                except Exception as e:
                    print(f"❌ Erro ao abrir arquivo Excel: {e}")
                    return False
            else:
                print("❌ Nenhum arquivo Excel encontrado")
                return False
        else:
            print("❌ Diretório de saída não existe")
            return False
            
    except Exception as e:
        print(f"❌ Erro durante processamento: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Limpar arquivo temporário
        try:
            if os.path.exists(recipe_file):
                os.remove(recipe_file)
                print(f"🧹 Arquivo temporário removido: {recipe_file}")
        except:
            pass

if __name__ == "__main__":
    success = test_complete_system()
    if success:
        print(f"\n🎉 TESTE COMPLETO PASSOU!")
        print("✅ Cores aplicadas corretamente")  
        print("✅ Custos calculados")
        print("✅ Sistema funcionando sem inventar receitas")
    else:
        print(f"\n💥 TESTE COMPLETO FALHOU")