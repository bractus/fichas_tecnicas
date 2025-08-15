import streamlit as st
import os
import json
from datetime import datetime
from dotenv import load_dotenv
from crewai import Agent, Task, Crew, Process
from crewai.tools import BaseTool
from langchain_openai import ChatOpenAI
from crewai_tools import SerperDevTool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# --- LÓGICA DO SCRIPT ORIGINAL (FERRAMENTAS, AGENTES, TAREFAS) ---
# Encapusulada para ser chamada pela interface

class ExcelGeneratorTool(BaseTool):
    name: str = "Gerador de Planilha de Ficha Técnica e Base de Insumos"
    description: str = "Cria um arquivo Excel a partir de um JSON com 'fichas_tecnicas' e 'base_de_insumos'."
    output_filename: str = "FICHA_TECNICA_GERADA.xlsx"

    def _run(self, dados_completos: dict) -> str:
        # Lógica de geração de Excel permanece a mesma...
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        fichas_tecnicas = dados_completos.get("fichas_tecnicas", [])
        base_de_insumos = dados_completos.get("base_de_insumos", [])
        font_titulo=Font(name='Arial', size=14, bold=True, color='FFFFFF')
        fill_titulo=PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        align_center=Alignment(horizontal='center', vertical='center')
        font_subtitulo=Font(name='Arial', size=11, bold=True)
        font_bold=Font(name='Arial', size=10, bold=True)
        fill_header_tabela=PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border_thin=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for ficha in fichas_tecnicas:
            sheet_name=ficha.get("nome_preparacao", "Ficha")[:31]
            sheet=workbook.create_sheet(title=sheet_name)
            sheet.merge_cells('A1:E1')
            cell=sheet['A1']
            cell.value="FICHA TÉCNICA DE PREPARAÇÃO"
            cell.font=font_titulo
            cell.fill=fill_titulo
            cell.alignment=align_center
            sheet.row_dimensions[1].height=25
            sheet['A3']="NOME DA PREPARAÇÃO:"
            sheet['A3'].font=font_subtitulo
            sheet.merge_cells('B3:E3')
            sheet['B3']=ficha.get("nome_preparacao", "").upper()
            sheet['B3'].font=font_subtitulo
            sheet['A4']="DEPARTAMENTO:"
            sheet['A4'].font=font_bold
            sheet['B4']="COZINHA"
            headers_tabela=["INGREDIENTES", "UNIDADE", "QUANTIDADE", "CUSTO UNITÁRIO", "CUSTO TOTAL"]
            for col_num, header_text in enumerate(headers_tabela, 1):
                cell=sheet.cell(row=7, column=col_num, value=header_text)
                cell.font=font_bold; cell.fill=fill_header_tabela; cell.border=border_thin
            linha_atual=8
            for ingrediente in ficha.get("ingredientes", []):
                sheet.cell(row=linha_atual, column=1, value=ingrediente[0])
                sheet.cell(row=linha_atual, column=2, value=ingrediente[1])
                sheet.cell(row=linha_atual, column=3, value=ingrediente[2])
                sheet.cell(row=linha_atual, column=4, value=ingrediente[3])
                sheet.cell(row=linha_atual, column=5, value=f'=C{linha_atual}*D{linha_atual}')
                for col in range(1, 6):
                    cell=sheet.cell(row=linha_atual, column=col)
                    cell.border=border_thin
                    if col==3: cell.number_format='0.000'
                    if col>=4: cell.number_format='R$ #,##0.00'
                linha_atual+=1
            primeira_linha_dados=8
            ultima_linha_dados=linha_atual - 1
            linha_totais=linha_atual + 1
            linha_seguinte=linha_totais
            if ultima_linha_dados>=primeira_linha_dados:
                sheet.cell(row=linha_totais, column=1, value="CUSTO TOTAL DA PREPARAÇÃO").font=font_bold
                custo_total_soma_cell=sheet.cell(row=linha_totais, column=5)
                custo_total_soma_cell.value=f'=SUM(E{primeira_linha_dados}:E{ultima_linha_dados})'
                custo_total_soma_cell.font=font_bold; custo_total_soma_cell.number_format='R$ #,##0.00'; custo_total_soma_cell.border=border_thin
                sheet.cell(row=linha_totais + 1, column=1, value="PESO TOTAL DA PREPARAÇÃO (KG)").font=font_bold
                peso_total_soma_cell=sheet.cell(row=linha_totais + 1, column=3)
                peso_total_soma_cell.value=f'=SUM(C{primeira_linha_dados}:C{ultima_linha_dados})'
                peso_total_soma_cell.font=font_bold; peso_total_soma_cell.number_format='0.000'; peso_total_soma_cell.border=border_thin
                sheet.cell(row=linha_totais + 2, column=1, value="CUSTO POR KG").font=font_bold
                custo_kg_cell=sheet.cell(row=linha_totais + 2, column=5)
                custo_kg_cell.value=f'={custo_total_soma_cell.coordinate}/{peso_total_soma_cell.coordinate}'
                custo_kg_cell.font=font_bold; custo_kg_cell.number_format='R$ #,##0.00'; custo_kg_cell.border=border_thin
                linha_seguinte=linha_totais + 3
                if "rendimento_porcoes" in ficha:
                    linha_adicionais=linha_seguinte + 1
                    rendimento=ficha["rendimento_porcoes"]
                    sheet.cell(row=linha_adicionais, column=1, value="RENDIMENTO (Nº DE PORÇÕES)").font=font_bold
                    sheet.cell(row=linha_adicionais, column=3, value=rendimento).font=font_bold
                    sheet.cell(row=linha_adicionais + 1, column=1, value="PESO POR PORÇÃO (KG)").font=font_bold
                    peso_porcao_cell=sheet.cell(row=linha_adicionais + 1, column=3)
                    peso_porcao_cell.value=f'={peso_total_soma_cell.coordinate}/{rendimento}'
                    peso_porcao_cell.font=font_bold; peso_porcao_cell.number_format='0.000'
                    sheet.cell(row=linha_adicionais + 2, column=1, value="CUSTO POR PORÇÃO").font=font_bold
                    custo_porcao_cell=sheet.cell(row=linha_adicionais + 2, column=5)
                    custo_porcao_cell.value=f'={custo_total_soma_cell.coordinate}/{rendimento}'
                    custo_porcao_cell.font=font_bold; custo_porcao_cell.number_format='R$ #,##0.00'
                    if "preco_venda" in ficha:
                        preco_venda=ficha["preco_venda"]
                        sheet.cell(row=linha_adicionais + 3, column=1, value="PREÇO DE VENDA").font=font_bold
                        preco_venda_cell=sheet.cell(row=linha_adicionais + 3, column=5, value=preco_venda)
                        preco_venda_cell.font=font_bold; preco_venda_cell.number_format='R$ #,##0.00'
                        sheet.cell(row=linha_adicionais + 4, column=1, value="CMV (%)").font=font_bold
                        cmv_cell=sheet.cell(row=linha_adicionais + 4, column=5)
                        cmv_cell.value=f'={custo_porcao_cell.coordinate}/{preco_venda_cell.coordinate}'
                        cmv_cell.font=font_bold; cmv_cell.number_format='0.00%'
                    linha_seguinte=linha_adicionais + 5
            if "modo_preparo" in ficha:
                linha_preparo=linha_seguinte + 2
                cell_titulo_preparo=sheet.cell(row=linha_preparo, column=1, value="MODO DE PREPARO")
                cell_titulo_preparo.font=font_subtitulo
                sheet.merge_cells(f'A{linha_preparo}:E{linha_preparo}')
                for i, passo in enumerate(ficha["modo_preparo"], 1):
                    passo_cell=sheet.cell(row=linha_preparo + i, column=1, value=passo)
                    passo_cell.alignment=Alignment(wrap_text=True, vertical='top')
                    sheet.merge_cells(f'A{linha_preparo + i}:E{linha_preparo + i}')
            sheet.column_dimensions['A'].width=35; sheet.column_dimensions['B'].width=12
            sheet.column_dimensions['C'].width=15; sheet.column_dimensions['D'].width=18
            sheet.column_dimensions['E'].width=18
        if base_de_insumos:
            sheet_base=workbook.create_sheet(title="Base de Insumos")
            headers_base=["INGREDIENTE", "UNIDADE", "PREÇO UNITÁRIO", "FORNECEDOR", "DATA DE COTAÇÃO"]
            for col, (header, width) in enumerate(zip(headers_base, [35, 15, 18, 25, 20]), 1):
                cell=sheet_base.cell(row=1, column=col, value=header)
                cell.font=font_bold; cell.fill=fill_header_tabela; cell.border=border_thin
                sheet_base.column_dimensions[get_column_letter(col)].width=width
            for row, item in enumerate(base_de_insumos, 2):
                sheet_base.cell(row=row, column=1, value=item.get("ingrediente"))
                sheet_base.cell(row=row, column=2, value=item.get("unidade"))
                preco_cell=sheet_base.cell(row=row, column=3, value=item.get("preco"))
                preco_cell.number_format='R$ #,##0.00'
                sheet_base.cell(row=row, column=4, value=item.get("fornecedor"))
                sheet_base.cell(row=row, column=5, value=item.get("data_cotacao"))
                for col in range(1, 6):
                    sheet_base.cell(row=row, column=col).border=border_thin
        
        workbook.save(self.output_filename)
        return f"Arquivo Excel '{self.output_filename}' criado com sucesso."

def get_crewai_setup(api_key: str):
    """Função para configurar e retornar a Crew do CrewAI."""
    load_dotenv()
    
    # 1. Ferramentas e Configuração do LLM
    os.environ["OPENAI_API_KEY"] = api_key
    llm = ChatOpenAI(model="gpt-4o", api_key=api_key, temperature=0.2)
    web_search = SerperDevTool()
    excel_tool = ExcelGeneratorTool()

    # 2. Agentes
    data_extractor_agent = Agent(role='Analista de Dados Culinários', goal='Analisar texto de receita, buscar preços online e extrair dados em JSON.', backstory='Mestre em NLP, capaz de extrair múltiplas fichas e encontrar informações na web.', tools=[web_search], verbose=True, llm=llm)
    excel_writer_agent = Agent(role='Gerador de Planilhas', goal='Receber um JSON e criar um arquivo Excel formatado com múltiplas abas.', backstory='Especialista em automação de planilhas que transforma dados brutos em relatórios Excel.', tools=[excel_tool], verbose=True, llm=llm)

    # 3. Tarefas
    task_extract_data = Task(
        agent=data_extractor_agent,
        description=f'''Analise o texto da receita fornecido abaixo e crie um ÚNICO objeto JSON com DUAS chaves: "fichas_tecnicas" e "base_de_insumos".

        **Texto da Receita para Análise:**
        ---
        {{recipe_text}}
        ---

        **Instruções Detalhadas:**
        1.  Para CADA ingrediente, se o preço não for fornecido, **use a busca na web para encontrar um preço de mercado realista em Reais (BRL)**.
        2.  A chave "fichas_tecnicas" deve ser uma LISTA de dicionários. Cada dicionário é uma receita com a estrutura:
            -   `nome_preparacao`: (string) O nome da receita.
            -   `rendimento_porcoes`: (number) O número de porções que a receita rende.
            -   `preco_venda`: (number) O preço de venda da porção.
            -   `ingredientes`: Uma LISTA de listas `[NOME, UNIDADE, QUANTIDADE, CUSTO_UNITARIO]`.
            -   `modo_preparo`: (Opcional) Uma LISTA de strings com os passos.
        3.  A chave "base_de_insumos" deve ser uma LISTA de dicionários. Cada dicionário é um insumo com a estrutura:
            -   `ingrediente`, `unidade`, `preco`, `fornecedor`, `data_cotacao` (hoje é {datetime.now().strftime('%Y-%m-%d')}).
        ''',
        expected_output='Um JSON único com as chaves "fichas_tecnicas" e "base_de_insumos", com preços pesquisados na web.'
    )

    task_generate_excel = Task(
        agent=excel_writer_agent,
        description='''Pegue o JSON completo da tarefa anterior. Use a ferramenta "Gerador de Planilha de Ficha Técnica e Base de Insumos" para criar o arquivo Excel final.''',
        context=[task_extract_data],
        expected_output='Mensagem de confirmação da criação do arquivo Excel.'
    )

    # 4. Montagem da Crew
    recipe_crew = Crew(
        agents=[data_extractor_agent, excel_writer_agent],
        tasks=[task_extract_data, task_generate_excel],
        process=Process.sequential, memory=False, verbose=True
    )
    
    return recipe_crew, excel_tool.output_filename


# --- INTERFACE STREAMLIT ---

st.set_page_config(page_title="Gerador de Ficha Técnica", layout="wide")
st.title("🍳 Gerador de Ficha Técnica com IA")
st.markdown("Cole sua receita, informe sua chave de API da OpenAI e deixe a IA fazer o resto!")

# Exemplo de receita para facilitar
example_recipe = """
FICHA TÉCNICA: FRANGO GRELHADO COM ARROZ

Rendimento (Nº de Porções): 1
Preço de Venda: 79.90

Ingredientes:
- Frango sassami: 180g
- Arroz: 150g
- Azeite: 10ml
- Sal: a gosto (considere 5g)

Modo de Preparo:
1. Tempere o frango com sal.
2. Aqueça o azeite em uma frigideira e grelhe o frango até dourar.
3. Cozinhe o arroz em água fervente.
4. Sirva o frango ao lado do arroz.
"""

# Colunas para layout
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Informe sua Receita")
    recipe_text = st.text_area("Cole a receita aqui:", value=example_recipe, height=400)

with col2:
    st.subheader("2. Configure e Execute")
    # Tenta carregar a chave do .env ou secrets do Streamlit, mas permite que o usuário insira
    api_key = st.text_input(
        "Sua Chave da API OpenAI:",
        type="password",
        help="Sua chave é usada apenas para esta sessão e não é armazenada.",
        value=os.getenv("OPENAI_API_KEY", "")
    )
    
    st.markdown("---")

    if st.button("✨ Gerar Ficha Técnica", use_container_width=True):
        if not recipe_text:
            st.error("Por favor, insira o texto da receita.")
        elif not api_key:
            st.error("Por favor, insira sua chave da API OpenAI.")
        else:
            try:
                with st.spinner("👩‍🍳 Processando... A IA está lendo a receita, buscando preços e montando sua planilha. Isso pode levar um minuto."):
                    # Configura e executa a Crew
                    recipe_crew, output_filename = get_crewai_setup(api_key)
                    inputs = {'recipe_text': recipe_text}
                    result = recipe_crew.kickoff(inputs=inputs)
                    
                    st.success("Ficha Técnica gerada com sucesso!")
                    st.markdown(f"**Resultado do processo:** `{result}`")

                    # Oferece o arquivo para download
                    with open(output_filename, "rb") as file:
                        st.download_button(
                            label=f"📥 Baixar {output_filename}",
                            data=file,
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

            except Exception as e:
                st.error(f"Ocorreu um erro durante o processo: {e}")
                st.error("Verifique se sua chave de API é válida e possui créditos. Se o erro persistir, a estrutura da receita pode ser muito complexa.")

st.markdown("---")
st.write("Desenvolvido com CrewAI e Streamlit.")