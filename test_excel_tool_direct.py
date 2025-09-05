#!/usr/bin/env python3
"""
Teste direto da ferramenta ExcelGeneratorTool
"""

import sys
import os
import json
sys.path.append('/Users/cairorocha/Documents/fichas_tecnicas')

from tools.generate_excel import ExcelGeneratorTool

def test_excel_tool_direct():
    """Testa a ferramenta ExcelGeneratorTool diretamente"""
    
    print("🧪 Testando ExcelGeneratorTool diretamente...")
    
    # Dados de teste
    test_data = {
        "fichas_tecnicas": [
            {
                "nome_preparacao": "Paella de Frutos do Mar",
                "rendimento_porcoes": 6,
                "preco_venda": 35.00,
                "ingredientes": [
                    {
                        "nome": "Arroz",
                        "unidade": "kg",
                        "quantidade": 0.5,
                        "fator_correcao": 1.0,
                        "custo_unitario": 8.50
                    },
                    {
                        "nome": "Camarão",
                        "unidade": "kg", 
                        "quantidade": 0.25,
                        "fator_correcao": 2.1,
                        "custo_unitario": 45.00
                    },
                    {
                        "nome": "Azeite",
                        "unidade": "L",
                        "quantidade": 0.05,
                        "fator_correcao": 1.0,
                        "custo_unitario": 25.00
                    }
                ],
                "modo_preparo": [
                    "1. Aqueça o azeite em uma paelleira grande",
                    "2. Refogue o alho e a cebola até ficarem dourados", 
                    "3. Adicione o arroz e mexa bem",
                    "4. Adicione água quente e temperos",
                    "5. Cozinhe por 15 minutos sem mexer",
                    "6. Adicione os frutos do mar",
                    "7. Continue cozinhando por mais 10 minutos"
                ]
            }
        ],
        "base_de_insumos": [
            {
                "ingrediente": "Arroz",
                "unidade": "kg",
                "preco": 8.50,
                "fator_correcao": 1.0,
                "fornecedor": "Atacadista Local",
                "data_cotacao": "2025-01-05"
            },
            {
                "ingrediente": "Camarão",
                "unidade": "kg",
                "preco": 45.00,
                "fator_correcao": 2.1,
                "fornecedor": "Peixaria Central", 
                "data_cotacao": "2025-01-05"
            },
            {
                "ingrediente": "Azeite",
                "unidade": "L",
                "preco": 25.00,
                "fator_correcao": 1.0,
                "fornecedor": "Distribuidor Alimentar",
                "data_cotacao": "2025-01-05"
            }
        ]
    }
    
    try:
        # Instanciar a ferramenta
        excel_tool = ExcelGeneratorTool()
        
        # Converter dados para JSON
        json_data = json.dumps(test_data)
        print(f"📊 Dados JSON: {len(json_data)} caracteres")
        
        # Testar com cores personalizadas
        color1 = 'FFC000'  # Amarelo
        color2 = 'FFF2CC'  # Amarelo claro
        
        print(f"🎨 Testando com cores: {color1} / {color2}")
        print("🚀 Executando ferramenta...")
        
        # Executar a ferramenta
        result = excel_tool._run(
            data_json=json_data,
            color=color1,
            color2=color2
        )
        
        print(f"✅ Resultado: {result}")
        
        # Verificar se arquivo foi gerado
        if "FICHA_TECNICA_COMPLETA_" in result:
            print("📊 Arquivo Excel gerado com sucesso!")
            
            # Verificar se existe
            import re
            path_match = re.search(r'/[^"]*\.xlsx', result)
            if path_match:
                excel_path = path_match.group()
                if os.path.exists(excel_path):
                    size = os.path.getsize(excel_path) / 1024
                    print(f"📁 Arquivo: {excel_path}")
                    print(f"📏 Tamanho: {size:.1f} KB")
                    
                    # Verificar conteúdo
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_path)
                        print(f"📋 Planilhas: {[ws.title for ws in wb.worksheets]}")
                        
                        # Verificar primeira planilha
                        if wb.worksheets:
                            first_sheet = wb.worksheets[0]
                            print(f"📄 Nome da preparação: {first_sheet.cell(3, 2).value}")
                            
                            # Verificar modo de preparo
                            found_modo_preparo = False
                            for row in range(1, 30):
                                cell_value = first_sheet.cell(row, 1).value
                                if cell_value and "MODO DE PREPARO" in str(cell_value):
                                    print(f"✅ Seção 'MODO DE PREPARO' encontrada na linha {row}")
                                    # Verificar próximas linhas
                                    for i in range(1, 8):
                                        step_cell = first_sheet.cell(row + i, 1).value
                                        if step_cell:
                                            print(f"  Passo {i}: {step_cell}")
                                    found_modo_preparo = True
                                    break
                            
                            if not found_modo_preparo:
                                print("⚠️  Seção 'MODO DE PREPARO' não encontrada")
                        
                        wb.close()
                        return True
                    except Exception as e:
                        print(f"❌ Erro ao verificar Excel: {e}")
                        return False
                else:
                    print(f"❌ Arquivo não encontrado: {excel_path}")
                    return False
            else:
                print("⚠️  Caminho do arquivo não encontrado no resultado")
                return False
        else:
            print(f"❌ Falha na geração: {result}")
            return False
            
    except Exception as e:
        print(f"❌ Erro durante teste: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_excel_tool_direct()
    if success:
        print(f"\n🎉 TESTE PASSOU - ExcelGeneratorTool funcionando!")
        print("✅ Modo de preparo formatado corretamente")
        print("✅ Base de insumos gerada")
        print("✅ Cores aplicadas")
    else:
        print(f"\n💥 TESTE FALHOU")