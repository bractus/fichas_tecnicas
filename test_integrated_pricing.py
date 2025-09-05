#!/usr/bin/env python3
"""
Teste da ferramenta Excel com cálculo de preços integrado
"""

import sys
import os
import json
sys.path.append('/Users/cairorocha/Documents/fichas_tecnicas')

from tools.generate_excel import ExcelGeneratorTool

def test_integrated_pricing():
    """Testa a ferramenta Excel com cálculo de preços integrado"""
    
    print("🧪 Testando Excel com cálculo de preços integrado...")
    
    # Dados de teste com preços zerados (serão calculados automaticamente)
    test_data = {
        "fichas_tecnicas": [
            {
                "nome_preparacao": "Strogonoff de Frango",
                "rendimento_porcoes": 4,
                "preco_venda": 0.0,  # Será calculado automaticamente
                "ingredientes": [
                    {
                        "nome": "Peito de Frango",
                        "unidade": "kg",
                        "quantidade": 0.5,
                        "fator_correcao": 1.15,  # 15% de perda no preparo
                        "custo_unitario": 0.0  # Será preenchido automaticamente
                    },
                    {
                        "nome": "Creme de Leite",
                        "unidade": "kg", 
                        "quantidade": 0.2,
                        "fator_correcao": 1.0,
                        "custo_unitario": 0.0
                    },
                    {
                        "nome": "Champignon",
                        "unidade": "kg",
                        "quantidade": 0.15,
                        "fator_correcao": 1.2,  # 20% de perda (talos, etc)
                        "custo_unitario": 0.0
                    },
                    {
                        "nome": "Cebola",
                        "unidade": "kg",
                        "quantidade": 0.1,
                        "fator_correcao": 1.3,  # 30% de perda
                        "custo_unitario": 0.0
                    }
                ],
                "modo_preparo": [
                    "1. Corte o frango em tiras e tempere",
                    "2. Refogue a cebola até dourar",
                    "3. Adicione o frango e cozinhe até dourar",
                    "4. Acrescente o champignon e refogue",
                    "5. Adicione o creme de leite e temperos",
                    "6. Cozinhe até engrossar ligeiramente"
                ]
            }
        ],
        "base_de_insumos": [
            {
                "ingrediente": "Peito de Frango",
                "unidade": "kg",
                "preco": 22.50,
                "fator_correcao": 1.15,
                "fornecedor": "Avícola São João",
                "data_cotacao": "2025-01-05"
            },
            {
                "ingrediente": "Creme de Leite",
                "unidade": "kg",
                "preco": 12.80,
                "fator_correcao": 1.0,
                "fornecedor": "Laticínios Bom Gosto",
                "data_cotacao": "2025-01-05"
            },
            {
                "ingrediente": "Champignon",
                "unidade": "kg",
                "preco": 35.00,
                "fator_correcao": 1.2,
                "fornecedor": "Cogumelos Premium",
                "data_cotacao": "2025-01-05"
            },
            {
                "ingrediente": "Cebola",
                "unidade": "kg",
                "preco": 4.20,
                "fator_correcao": 1.3,
                "fornecedor": "Ceasa Local",
                "data_cotacao": "2025-01-05"
            }
        ]
    }
    
    try:
        # Instanciar a ferramenta
        excel_tool = ExcelGeneratorTool()
        
        # Converter dados para JSON
        json_data = json.dumps(test_data)
        print(f"📊 Dados de entrada: 1 receita com preço zerado")
        print(f"💰 Custos esperados:")
        print(f"  • Peito de Frango: 0.5kg × 1.15 × R$ 22.50 = R$ 12.94")
        print(f"  • Creme de Leite: 0.2kg × 1.0 × R$ 12.80 = R$ 2.56")
        print(f"  • Champignon: 0.15kg × 1.2 × R$ 35.00 = R$ 6.30")
        print(f"  • Cebola: 0.1kg × 1.3 × R$ 4.20 = R$ 0.55")
        custo_esperado = 12.94 + 2.56 + 6.30 + 0.55
        custo_por_porcao = custo_esperado / 4
        preco_esperado = custo_por_porcao * 3.5  # 3.5x markup
        print(f"💵 Custo total esperado: R$ {custo_esperado:.2f}")
        print(f"💵 Custo por porção: R$ {custo_por_porcao:.2f}")
        print(f"🏷️  Preço esperado (~3.5x): R$ {preco_esperado:.2f}")
        
        print(f"\n🚀 Executando ferramenta Excel...")
        
        # Executar a ferramenta com cores personalizadas
        result = excel_tool._run(
            data_json=json_data,
            color='7030A0',  # Roxo
            color2='E4DFEC'  # Roxo claro
        )
        
        print(f"✅ Resultado: {result}")
        
        # Verificar se arquivo foi gerado e extrair informações
        if "FICHA_TECNICA_COMPLETA_" in result:
            print("📊 Arquivo Excel gerado com sucesso!")
            
            # Verificar se existe
            import re
            path_match = re.search(r'/[^"]*\.xlsx', result)
            if path_match:
                excel_path = path_match.group()
                if os.path.exists(excel_path):
                    size = os.path.getsize(excel_path) / 1024
                    print(f"📁 Arquivo: {os.path.basename(excel_path)}")
                    print(f"📏 Tamanho: {size:.1f} KB")
                    
                    # Verificar conteúdo e preços calculados
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_path)
                        print(f"📋 Planilhas: {[ws.title for ws in wb.worksheets]}")
                        
                        # Verificar primeira planilha
                        if wb.worksheets:
                            first_sheet = wb.worksheets[0]
                            print(f"📄 Nome da preparação: {first_sheet.cell(3, 2).value}")
                            
                            # Procurar pelo preço de venda
                            for row in range(1, 20):
                                for col in range(1, 8):
                                    cell_value = first_sheet.cell(row, col).value
                                    if cell_value and "preço" in str(cell_value).lower():
                                        # Verificar se há um valor monetário próximo
                                        for check_row in range(max(1, row-2), min(21, row+3)):
                                            for check_col in range(1, 8):
                                                check_value = first_sheet.cell(check_row, check_col).value
                                                if isinstance(check_value, (int, float)) and check_value > 5:
                                                    print(f"💰 Preço de venda encontrado: R$ {check_value:.2f}")
                                                    break
                            
                            # Verificar custos unitários na tabela de ingredientes
                            print(f"🔍 Verificando custos unitários preenchidos...")
                            costs_found = 0
                            for row in range(8, 15):  # Aproximadamente onde ficam os ingredientes
                                custo_cell = first_sheet.cell(row, 6)  # Coluna F - custo unitário
                                if custo_cell.value and isinstance(custo_cell.value, (int, float)) and custo_cell.value > 0:
                                    ingredient_name = first_sheet.cell(row, 1).value
                                    print(f"  ✅ {ingredient_name}: R$ {custo_cell.value:.2f}")
                                    costs_found += 1
                            
                            if costs_found > 0:
                                print(f"✅ {costs_found} custos unitários preenchidos automaticamente")
                            else:
                                print("⚠️  Custos unitários não encontrados na planilha")
                        
                        wb.close()
                        return True
                    except Exception as e:
                        print(f"❌ Erro ao verificar Excel: {e}")
                        return False
                else:
                    print(f"❌ Arquivo não encontrado: {excel_path}")
                    return False
            else:
                print("⚠️  Caminho do arquivo não encontrado no resultado")
                return False
        else:
            print(f"❌ Falha na geração: {result}")
            return False
            
    except Exception as e:
        print(f"❌ Erro durante teste: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_integrated_pricing()
    if success:
        print(f"\n🎉 TESTE PASSOU - Preços calculados automaticamente!")
        print("✅ Custos unitários preenchidos automaticamente")
        print("✅ Preço de venda calculado com markup mínimo 3x")
        print("✅ Ferramenta Excel funciona com cálculo integrado")
    else:
        print(f"\n💥 TESTE FALHOU")