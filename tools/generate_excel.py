from crewai.tools import BaseTool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import json
import logging
import re
from pydantic import BaseModel, Field
from typing import Type, Dict, List, Any

logger = logging.getLogger(__name__)

class ExcelGeneratorInput(BaseModel):
    data_json: str = Field(..., description="JSON data containing fichas_tecnicas and base_de_insumos to generate Excel file")
    color: str = Field(default='4472C4', description="Primary color for Excel headers (hexadecimal without #)")
    color2: str = Field(default='D9E1F2', description="Secondary color for Excel tables (hexadecimal without #)")

class ExcelGeneratorTool(BaseTool):
    name: str = "Gerador de Planilha de Ficha Técnica e Base de Insumos"
    description: str = "Creates Excel file with one sheet per technical sheet and ONE unified ingredient base sheet. Provide data as JSON string."
    args_schema: Type[BaseModel] = ExcelGeneratorInput

    def _calculate_viable_price(self, ficha: dict, insumos_dict: dict) -> None:
        """Calcula preço de venda viável baseado no custo dos ingredientes (mínimo 3x)"""
        import math
        
        nome_receita = ficha.get('nome_preparacao', 'Receita')
        rendimento = ficha.get('rendimento_porcoes', 1)
        
        total_cost = 0.0
        
        # Calcular custo total da receita
        for ingrediente in ficha.get('ingredientes', []):
            nome = ingrediente.get('nome', '').lower()
            quantidade = ingrediente.get('quantidade', 0)
            fator_correcao = ingrediente.get('fator_correcao', 1.0)
            
            # Buscar preço no dicionário de insumos e atualizar custo_unitario
            if nome in insumos_dict:
                preco_unitario = insumos_dict[nome].get('preco', 0.0)
                ingrediente['custo_unitario'] = preco_unitario
                
                # Calcular custo do ingrediente
                quantidade_corrigida = quantidade * fator_correcao
                custo_ingrediente = quantidade_corrigida * preco_unitario
                total_cost += custo_ingrediente
        
        # Calcular preço de venda
        if total_cost > 0 and rendimento > 0:
            custo_por_porcao = total_cost / rendimento
            
            # Aplicar markup de 3.5x (mínimo 3x)
            preco_base = custo_por_porcao * 3.5
            
            # Arredondar para o próximo 0.50
            preco_venda = math.ceil(preco_base * 2) / 2
            
            # Garantir mínimo de 3x
            preco_minimo = custo_por_porcao * 3
            if preco_venda < preco_minimo:
                preco_venda = math.ceil(preco_minimo * 2) / 2
            
            # Atualizar preço de venda na ficha
            ficha['preco_venda'] = preco_venda
            
            markup = (preco_venda / custo_por_porcao) if custo_por_porcao > 0 else 0
            logger.info(f"💰 {nome_receita}: Custo R$ {total_cost:.2f} → Preço R$ {preco_venda:.2f}/porção (markup {markup:.1f}x)")
        else:
            # Se não conseguir calcular, usar preço padrão
            if ficha.get('preco_venda', 0) <= 0:
                ficha['preco_venda'] = 15.0
            logger.warning(f"⚠️  {nome_receita}: Não foi possível calcular preço, usando padrão R$ {ficha['preco_venda']:.2f}")

    def _run(self, data_json: str, color: str = '4472C4', color2: str = 'D9E1F2') -> str:
        """Cria um arquivo Excel profissional a partir de dados JSON"""
        filepath = None
        try:
            logger.info("Starting Excel generation with ExcelGeneratorTool")
            current_working_dir = os.getcwd()
            logger.info(f"Current working directory in ExcelGeneratorTool: {current_working_dir}")
            
            # Use dynamic project directory (works on Railway and local)
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            output_dir = os.path.join(project_root, "output")
            
            logger.info(f"Current working directory: {current_working_dir}")
            logger.info(f"Project root directory: {project_root}")
            logger.info(f"Using output directory: {output_dir}")
            
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            # Clean up any existing invalid Excel files first
            self._cleanup_invalid_excel_files(output_dir)
            
            # Parse JSON se for string
            if isinstance(data_json, str):
                try:
                    data = json.loads(data_json)
                except json.JSONDecodeError:
                    # Try to extract JSON from text
                    json_match = re.search(r'\{.*\}', data_json, re.DOTALL)
                    if json_match:
                        try:
                            data = json.loads(json_match.group())
                        except json.JSONDecodeError:
                            logger.error(f"Failed to parse JSON: {data_json[:100]}...")
                            return "❌ ERRO: Não foi possível extrair JSON válido dos dados fornecidos."
                    else:
                        logger.error("No JSON found in input data")
                        return "❌ ERRO: Não foi possível extrair JSON válido dos dados fornecidos."
            else:
                data = data_json
            
            # Create workbook
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            fichas = data.get('fichas_tecnicas', [])
            insumos = data.get('base_de_insumos', [])
            
            logger.info(f"Processing {len(fichas)} fichas and {len(insumos)} insumos")
            
            # Create insumos lookup dictionary for price calculation
            insumos_dict = {}
            for insumo in insumos:
                nome_key = insumo.get('ingrediente', '').lower()
                insumos_dict[nome_key] = insumo
            
            # Calculate viable selling prices for each recipe
            for ficha in fichas:
                self._calculate_viable_price(ficha, insumos_dict)
            
            # Debug: Log detailed information about the data received including modo_preparo
            if fichas:
                logger.info("=== FICHAS TÉCNICAS RECEBIDAS ===")
                for i, ficha in enumerate(fichas, 1):
                    nome = ficha.get('nome_preparacao', 'NOME NÃO ENCONTRADO')
                    ingredientes_count = len(ficha.get('ingredientes', []))
                    modo_preparo = ficha.get('modo_preparo', [])
                    modo_preparo_status = f"{len(modo_preparo)} passos" if modo_preparo else "AUSENTE"
                    logger.info(f"Ficha {i}: {nome} - {ingredientes_count} ingredientes - Modo de Preparo: {modo_preparo_status}")
                    if modo_preparo:
                        logger.info(f"  Passos de preparo: {modo_preparo[:3]}{'...' if len(modo_preparo) > 3 else ''}")
                    else:
                        logger.warning(f"  ⚠️  MODO DE PREPARO AUSENTE para: {nome}")
            else:
                logger.warning("NENHUMA FICHA TÉCNICA ENCONTRADA NO JSON!")
                
            if insumos:
                logger.info("=== INSUMOS RECEBIDOS ===")
                for i, insumo in enumerate(insumos, 1):
                    ingrediente = insumo.get('ingrediente', 'INGREDIENTE NÃO ENCONTRADO')
                    preco = insumo.get('preco', 'PREÇO NÃO ENCONTRADO')
                    logger.info(f"Insumo {i}: {ingrediente} - R$ {preco}")
            else:
                logger.warning("NENHUM INSUMO ENCONTRADO NO JSON!")
                
            # Log the raw JSON structure keys for debugging
            logger.info(f"JSON keys received: {list(data.keys())}")
            if 'fichas_tecnicas' in data:
                logger.info(f"fichas_tecnicas type: {type(data['fichas_tecnicas'])}")
            if 'base_de_insumos' in data:
                logger.info(f"base_de_insumos type: {type(data['base_de_insumos'])}")

            # Estilos profissionais (copiados da ExcelGeneratorTool)
            font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            fill_titulo = PatternFill(start_color=color, end_color=color, fill_type='solid')
            align_center = Alignment(horizontal='center', vertical='center')
            font_subtitulo = Font(name='Arial', size=11, bold=True)
            font_bold = Font(name='Arial', size=10, bold=True)
            fill_header_tabela = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # --- PARTE 1: Gerar uma aba para cada Ficha Técnica ---
            for idx, ficha in enumerate(fichas, 1):
                sheet_name = f"Ficha_{idx}_{ficha.get('nome_preparacao', 'Receita')[:20]}"
                sheet = workbook.create_sheet(title=sheet_name)

                # A. Cabeçalho Principal e Informações
                sheet.merge_cells('A1:G1')
                cell = sheet['A1']
                cell.value = "FICHA TÉCNICA DE PREPARAÇÃO"
                cell.font = font_titulo
                cell.fill = fill_titulo
                cell.alignment = align_center
                sheet.row_dimensions[1].height = 25
                
                sheet['A3'] = "NOME DA PREPARAÇÃO:"
                sheet['A3'].font = font_subtitulo
                sheet.merge_cells('B3:G3')
                sheet['B3'] = ficha.get("nome_preparacao", "").upper()
                sheet['B3'].font = font_subtitulo
                
                sheet['A4'] = "DEPARTAMENTO:"
                sheet['A4'].font = font_bold
                sheet['B4'] = "COZINHA"

                # B. Tabela de Ingredientes
                headers_tabela = ["INGREDIENTES", "UNIDADE", "QUANTIDADE", "FATOR CORREÇÃO", "PESO CORRIGIDO", "CUSTO UNITÁRIO", "CUSTO TOTAL"]
                for col_num, header_text in enumerate(headers_tabela, 1):
                    cell = sheet.cell(row=7, column=col_num, value=header_text)
                    cell.font = font_bold
                    cell.fill = fill_header_tabela
                    cell.border = border_thin
                
                linha_atual = 8
                for ingrediente in ficha.get("ingredientes", []):
                    # Handle both list and dict formats
                    if isinstance(ingrediente, list) and len(ingrediente) >= 5:
                        nome, unidade, quantidade_original, fator_correcao, custo_unitario = ingrediente[:5]
                        quantidade_corrigida = quantidade_original * fator_correcao
                    elif isinstance(ingrediente, dict):
                        nome = ingrediente.get('nome', 'N/A')
                        unidade = ingrediente.get('unidade', 'N/A')
                        quantidade_original = ingrediente.get('quantidade', 0)
                        fator_correcao = ingrediente.get('fator_correcao', 1.0)
                        custo_unitario = ingrediente.get('custo_unitario', 0)
                        quantidade_corrigida = quantidade_original * fator_correcao
                    else:
                        continue

                    sheet.cell(row=linha_atual, column=1, value=nome)
                    sheet.cell(row=linha_atual, column=2, value=unidade)
                    sheet.cell(row=linha_atual, column=3, value=quantidade_original)
                    sheet.cell(row=linha_atual, column=4, value=fator_correcao)
                    sheet.cell(row=linha_atual, column=5, value=quantidade_corrigida)
                    sheet.cell(row=linha_atual, column=6, value=custo_unitario)
                    sheet.cell(row=linha_atual, column=7, value=f'=E{linha_atual}*F{linha_atual}')  # Peso corrigido × custo
                    
                    for col in range(1, 8):
                        cell = sheet.cell(row=linha_atual, column=col)
                        cell.border = border_thin
                        if col in [3, 4, 5]:  # Quantidade, fator, peso corrigido
                            cell.number_format = '0.000'
                        if col in [6, 7]:  # Custos
                            cell.number_format = 'R$ #,##0.00'
                    linha_atual += 1

                # C. Seção de Totais e Fórmulas
                primeira_linha_dados = 8
                ultima_linha_dados = linha_atual - 1
                linha_totais = linha_atual + 1
                linha_seguinte = linha_totais

                if ultima_linha_dados >= primeira_linha_dados:
                    # Custo Total da Preparação
                    sheet.cell(row=linha_totais, column=1, value="CUSTO TOTAL DA PREPARAÇÃO").font = font_bold
                    custo_total_soma_cell = sheet.cell(row=linha_totais, column=7)  # Column G (custo total)
                    custo_total_soma_cell.value = f'=SUM(G{primeira_linha_dados}:G{ultima_linha_dados})'
                    custo_total_soma_cell.font = font_bold
                    custo_total_soma_cell.number_format = 'R$ #,##0.00'
                    custo_total_soma_cell.border = border_thin

                    # Peso Total da Preparação (usando peso corrigido)
                    sheet.cell(row=linha_totais + 1, column=1, value="PESO TOTAL DA PREPARAÇÃO (KG)").font = font_bold
                    peso_total_soma_cell = sheet.cell(row=linha_totais + 1, column=5)  # Column E (peso corrigido)
                    peso_total_soma_cell.value = f'=SUM(E{primeira_linha_dados}:E{ultima_linha_dados})'
                    peso_total_soma_cell.font = font_bold
                    peso_total_soma_cell.number_format = '0.000'
                    peso_total_soma_cell.border = border_thin

                    # Custo por Kg
                    sheet.cell(row=linha_totais + 2, column=1, value="CUSTO POR KG").font = font_bold
                    custo_kg_cell = sheet.cell(row=linha_totais + 2, column=7)
                    custo_kg_cell.value = f'={custo_total_soma_cell.coordinate}/{peso_total_soma_cell.coordinate}'
                    custo_kg_cell.font = font_bold
                    custo_kg_cell.number_format = 'R$ #,##0.00'
                    custo_kg_cell.border = border_thin
                    
                    linha_seguinte = linha_totais + 3

                    # D. Seção de Rendimento e Preço
                    if "rendimento_porcoes" in ficha:
                        linha_adicionais = linha_seguinte + 1
                        rendimento = ficha["rendimento_porcoes"]
                        
                        sheet.cell(row=linha_adicionais, column=1, value="RENDIMENTO (Nº DE PORÇÕES)").font = font_bold
                        sheet.cell(row=linha_adicionais, column=3, value=rendimento).font = font_bold
                        
                        sheet.cell(row=linha_adicionais + 1, column=1, value="PESO POR PORÇÃO (KG)").font = font_bold
                        peso_porcao_cell = sheet.cell(row=linha_adicionais + 1, column=5)  # Column E
                        peso_porcao_cell.value = f'={peso_total_soma_cell.coordinate}/{rendimento}'
                        peso_porcao_cell.font = font_bold
                        peso_porcao_cell.number_format = '0.000'
                        
                        sheet.cell(row=linha_adicionais + 2, column=1, value="CUSTO POR PORÇÃO").font = font_bold
                        custo_porcao_cell = sheet.cell(row=linha_adicionais + 2, column=7)  # Column G
                        custo_porcao_cell.value = f'={custo_total_soma_cell.coordinate}/{rendimento}'
                        custo_porcao_cell.font = font_bold
                        custo_porcao_cell.number_format = 'R$ #,##0.00'

                        if "preco_venda" in ficha:
                            preco_venda = ficha["preco_venda"]
                            sheet.cell(row=linha_adicionais + 3, column=1, value="PREÇO DE VENDA").font = font_bold
                            preco_venda_cell = sheet.cell(row=linha_adicionais + 3, column=7, value=preco_venda)  # Column G
                            preco_venda_cell.font = font_bold
                            preco_venda_cell.number_format = 'R$ #,##0.00'

                            sheet.cell(row=linha_adicionais + 4, column=1, value="CMV (%)").font = font_bold
                            cmv_cell = sheet.cell(row=linha_adicionais + 4, column=7)  # Column G
                            cmv_cell.value = f'={custo_porcao_cell.coordinate}/{preco_venda_cell.coordinate}'
                            cmv_cell.font = font_bold
                            cmv_cell.number_format = '0.00%'

                        linha_seguinte = linha_adicionais + 5

                # E. Seção Modo de Preparo (SEMPRE INCLUÍDA)
                modo_preparo = ficha.get("modo_preparo", [])
                logger.info(f"Processing modo_preparo for {ficha.get('nome_preparacao', 'Unknown')}: {len(modo_preparo) if modo_preparo else 0} steps")
                
                # SEMPRE adicionar seção de modo de preparo
                linha_preparo = linha_seguinte + 2
                cell_titulo_preparo = sheet.cell(row=linha_preparo, column=1, value="MODO DE PREPARO")
                cell_titulo_preparo.font = font_subtitulo
                sheet.merge_cells(f'A{linha_preparo}:G{linha_preparo}')
                
                if "modo_preparo" in ficha and ficha["modo_preparo"]:
                    # Modo de preparo encontrado - adicionar passos
                    modo_preparo_data = ficha["modo_preparo"]
                    logger.info(f"MODO DE PREPARO DEBUG: type={type(modo_preparo_data)}, value={modo_preparo_data}")
                    
                    # Verificar se é string ao invés de lista
                    if isinstance(modo_preparo_data, str):
                        logger.warning("modo_preparo is string, converting to list")
                        # Se for string, tenta quebrar por \n ou por numeração
                        if '\n' in modo_preparo_data:
                            steps = [step.strip() for step in modo_preparo_data.split('\n') if step.strip()]
                        else:
                            steps = [modo_preparo_data]  # Uma única string
                    elif isinstance(modo_preparo_data, list):
                        steps = modo_preparo_data
                    else:
                        logger.error(f"Unknown modo_preparo format: {type(modo_preparo_data)}")
                        steps = ["Modo de preparo inválido"]
                    
                    logger.info(f"Adding MODO DE PREPARO section with {len(steps)} steps")
                    for i, passo in enumerate(steps, 1):
                        logger.info(f"Adding step {i}: {passo[:50]}{'...' if len(str(passo)) > 50 else ''}")
                        passo_cell = sheet.cell(row=linha_preparo + i, column=1, value=str(passo))
                        passo_cell.alignment = Alignment(wrap_text=True, vertical='top')
                        sheet.merge_cells(f'A{linha_preparo + i}:G{linha_preparo + i}')
                        sheet.row_dimensions[linha_preparo + i].height = 30
                        linha_seguinte = linha_preparo + i
                else:
                    # Modo de preparo vazio - adicionar mensagem padrão
                    logger.warning(f"⚠️  modo_preparo is empty for: {ficha.get('nome_preparacao', 'Unknown')} - adding default message")
                    passo_cell = sheet.cell(row=linha_preparo + 1, column=1, value="Modo de preparo não especificado")
                    passo_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                    passo_cell.font = Font(italic=True, color="666666")
                    sheet.merge_cells(f'A{linha_preparo + 1}:G{linha_preparo + 1}')
                    sheet.row_dimensions[linha_preparo + 1].height = 25
                    linha_seguinte = linha_preparo + 1

                # F. Ajuste da Largura das Colunas
                sheet.column_dimensions['A'].width = 30  # Ingredientes
                sheet.column_dimensions['B'].width = 10  # Unidade
                sheet.column_dimensions['C'].width = 12  # Quantidade
                sheet.column_dimensions['D'].width = 12  # Fator Correção
                sheet.column_dimensions['E'].width = 15  # Peso Corrigido
                sheet.column_dimensions['F'].width = 15  # Custo Unitário
                sheet.column_dimensions['G'].width = 15  # Custo Total
            
            # --- PARTE 2: Gerar UMA ÚNICA aba "Base de Insumos Unificada" ---
            logger.info(f"BASE DE INSUMOS DEBUG: Found {len(insumos)} insumos")
            if insumos:
                logger.info("Creating Base de Insumos sheet")
                for i, insumo in enumerate(insumos[:3], 1):  # Log first 3
                    logger.info(f"Insumo {i}: {insumo}")
                sheet_base = workbook.create_sheet(title="Base de Insumos")
                
                # Cabeçalho
                headers_base = ["INGREDIENTE", "UNIDADE", "PREÇO UNITÁRIO", "FATOR CORREÇÃO", "FORNECEDOR", "DATA DE COTAÇÃO"]
                widths = [30, 12, 15, 15, 25, 18]
                
                for col, (header, width) in enumerate(zip(headers_base, widths), 1):
                    cell = sheet_base.cell(row=1, column=col, value=header)
                    cell.font = font_bold
                    cell.fill = fill_header_tabela
                    cell.border = border_thin
                    sheet_base.column_dimensions[get_column_letter(col)].width = width
                
                # Dados dos insumos
                for row, item in enumerate(insumos, 2):
                    sheet_base.cell(row=row, column=1, value=item.get("ingrediente"))
                    sheet_base.cell(row=row, column=2, value=item.get("unidade"))
                    
                    preco_cell = sheet_base.cell(row=row, column=3, value=item.get("preco"))
                    preco_cell.number_format = 'R$ #,##0.00'
                    
                    fator_cell = sheet_base.cell(row=row, column=4, value=item.get("fator_correcao", 1.0))
                    fator_cell.number_format = '0.000'
                    
                    sheet_base.cell(row=row, column=5, value=item.get("fornecedor"))
                    sheet_base.cell(row=row, column=6, value=item.get("data_cotacao"))
                    
                    for col in range(1, 7):
                        sheet_base.cell(row=row, column=col).border = border_thin
            else:
                logger.warning("No insumos data found - creating empty Base de Insumos sheet")
                sheet_base = workbook.create_sheet(title="Base de Insumos")
                
                # Cabeçalho
                headers_base = ["INGREDIENTE", "UNIDADE", "PREÇO UNITÁRIO", "FATOR CORREÇÃO", "FORNECEDOR", "DATA DE COTAÇÃO"]
                widths = [30, 12, 15, 15, 25, 18]
                
                for col, (header, width) in enumerate(zip(headers_base, widths), 1):
                    cell = sheet_base.cell(row=1, column=col, value=header)
                    cell.font = font_bold
                    cell.fill = fill_header_tabela
                    cell.border = border_thin
                    sheet_base.column_dimensions[get_column_letter(col)].width = width
                
                # Mensagem informativa
                no_data_cell = sheet_base.cell(row=2, column=1, value="Nenhum insumo gerado - verifique o processamento")
                no_data_cell.font = Font(italic=True, color="999999")
                sheet_base.merge_cells('A2:F2')
            
            # Generate unique filename to prevent conflicts
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"FICHA_TECNICA_COMPLETA_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Save workbook with error handling
            try:
                workbook.save(filepath)
                logger.info(f"Workbook saved to: {filepath}")
                
                # Verify file was created and is valid
                if self._verify_excel_file(filepath):
                    logger.info(f"Excel file created and verified successfully: {filepath}")
                    
                    return f"✅ SUCESSO: Arquivo Excel criado: {filepath} com {len(fichas)} receitas e {len(insumos)} insumos."
                else:
                    # Remove invalid file if it exists
                    if os.path.exists(filepath):
                        os.remove(filepath)
                        logger.warning(f"Removed invalid Excel file: {filepath}")
                    return "❌ ERRO: Arquivo Excel não foi criado corretamente."
                    
            except Exception as save_error:
                logger.error(f"Error saving Excel file: {save_error}")
                # Clean up any partial file
                if filepath and os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                        logger.info(f"Cleaned up partial file: {filepath}")
                    except:
                        pass
                return f"❌ ERRO ao salvar arquivo Excel: {str(save_error)}"
                
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            # Clean up any partial file
            if filepath and os.path.exists(filepath):
                try:
                    os.remove(filepath)
                    logger.info(f"Cleaned up partial file after error: {filepath}")
                except:
                    pass
            return f"❌ ERRO na criação do Excel: {str(e)}"
    
    def _cleanup_invalid_excel_files(self, output_dir: str) -> None:
        """Remove arquivos Excel inválidos ou corrompidos do diretório"""
        try:
            if not os.path.exists(output_dir):
                return
                
            excel_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
            for filename in excel_files:
                filepath = os.path.join(output_dir, filename)
                try:
                    # Try to verify if file is valid
                    if not self._verify_excel_file(filepath):
                        os.remove(filepath)
                        logger.info(f"Removed invalid Excel file: {filepath}")
                except Exception as e:
                    # If we can't even check the file, it's probably corrupted
                    try:
                        os.remove(filepath)
                        logger.info(f"Removed corrupted Excel file: {filepath}")
                    except:
                        logger.warning(f"Could not remove file: {filepath}")
        except Exception as e:
            logger.warning(f"Error during cleanup: {e}")
    
    def _verify_excel_file(self, filepath: str) -> bool:
        """Verifica se o arquivo Excel é válido"""
        try:
            if not os.path.exists(filepath):
                return False
                
            # Check file size (should be > 0 and reasonable minimum)
            file_size = os.path.getsize(filepath)
            if file_size < 1024:  # Less than 1KB is probably invalid
                return False
                
            # Try to open with openpyxl to verify structure
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            
            # Basic validation - should have at least one worksheet
            if len(wb.worksheets) == 0:
                return False
                
            # Check if we can access the first worksheet
            first_sheet = wb.worksheets[0]
            
            # Try to read a cell to ensure the file isn't corrupted
            _ = first_sheet.cell(1, 1).value
            
            wb.close()
            return True
            
        except Exception as e:
            logger.warning(f"File verification failed for {filepath}: {e}")
            return False
